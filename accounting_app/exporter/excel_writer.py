import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, Reference
from typing import Dict, List, Any, Optional
import os
from datetime import datetime

class ExcelWriter:
    def __init__(self):
        self.styles = self._define_styles()
    
    def _define_styles(self) -> Dict[str, Any]:
        """Define Excel styles for professional formatting"""
        # Define borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        return {
            'header': {
                'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
                'font': Font(color="FFFFFF", bold=True, size=12),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': thin_border
            },
            'subheader': {
                'fill': PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid"),
                'font': Font(color="FFFFFF", bold=True, size=11),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': thin_border
            },
            'amount': {
                'font': Font(size=10),
                'alignment': Alignment(horizontal='right', vertical='center'),
                'number_format': '#,##0.00',
                'border': thin_border
            },
            'date': {
                'font': Font(size=10),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'number_format': 'dd/mm/yyyy',
                'border': thin_border
            },
            'text': {
                'font': Font(size=10),
                'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
                'border': thin_border
            },
            'debit': {
                'font': Font(size=10, color="FF0000"),
                'alignment': Alignment(horizontal='right', vertical='center'),
                'number_format': '#,##0.00',
                'border': thin_border
            },
            'credit': {
                'font': Font(size=10, color="008000"),
                'alignment': Alignment(horizontal='right', vertical='center'),
                'number_format': '#,##0.00',
                'border': thin_border
            },
            'total': {
                'fill': PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
                'font': Font(bold=True, size=11),
                'alignment': Alignment(horizontal='right', vertical='center'),
                'number_format': '#,##0.00',
                'border': thin_border
            }
        }
    
    def export_to_excel(self, transactions_df: pd.DataFrame, journal_entries: List[Dict[str, Any]], 
                       output_path: str, include_charts: bool = True) -> str:
        """
        Export transactions and journal entries to Excel with professional formatting
        
        Args:
            transactions_df: DataFrame with transactions
            journal_entries: List of journal entries
            output_path: Output file path
            include_charts: Whether to include charts and analysis
            
        Returns:
            Path to created file
        """
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                workbook = writer.book
                
                # Create worksheets
                self._create_transactions_sheet(workbook, transactions_df, writer)
                self._create_journal_sheet(workbook, journal_entries, writer)
                self._create_summary_sheet(workbook, transactions_df, journal_entries, writer)
                
                if include_charts:
                    self._create_analysis_sheet(workbook, transactions_df, journal_entries, writer)
                
                # Add formatting
                self._apply_auto_filters(workbook)
                
            print(f"Excel file exported successfully: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            raise
    
    def _create_transactions_sheet(self, workbook: Workbook, transactions_df: pd.DataFrame, writer):
        """Create transactions worksheet"""
        if transactions_df.empty:
            return
        
        # Create transactions sheet
        sheet = workbook.create_sheet("Transactions", 0)
        
        # Add title
        sheet.merge_cells('A1:H1')
        sheet['A1'] = "BANK TRANSACTIONS SUMMARY"
        sheet['A1'].font = Font(size=16, bold=True)
        sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Prepare data for export
        export_df = transactions_df.copy()
        
        # Convert date objects to strings for better Excel handling
        if 'date' in export_df.columns:
            export_df['date'] = export_df['date'].apply(
                lambda x: x.strftime('%d/%m/%Y') if hasattr(x, 'strftime') else str(x)
            )
        
        # Write data starting from row 3
        rows = dataframe_to_rows(export_df, index=False, header=True)
        for r_idx, row in enumerate(rows, 3):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
        
        # Apply formatting
        self._format_data_range(sheet, 3, 1, len(export_df) + 3, len(export_df.columns))
        
        # Add headers
        self._apply_header_formatting(sheet, 3, export_df.columns.tolist())
        
        # Auto-adjust column widths
        self._auto_adjust_columns(sheet, export_df)
        
        # Add totals row
        self._add_totals_row(sheet, export_df, len(export_df) + 4)
    
    def _create_journal_sheet(self, workbook: Workbook, journal_entries: List[Dict[str, Any]], writer):
        """Create journal entries worksheet"""
        if not journal_entries:
            return
        
        # Convert to DataFrame
        journal_df = pd.DataFrame(journal_entries)
        
        # Create journal sheet
        sheet = workbook.create_sheet("Journal Entries", 1)
        
        # Add title
        sheet.merge_cells('A1:G1')
        sheet['A1'] = "ACCOUNTING JOURNAL ENTRIES"
        sheet['A1'].font = Font(size=16, bold=True)
        sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Write data starting from row 3
        rows = dataframe_to_rows(journal_df, index=False, header=True)
        for r_idx, row in enumerate(rows, 3):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
        
        # Apply formatting
        self._format_data_range(sheet, 3, 1, len(journal_df) + 3, len(journal_df.columns))
        self._apply_header_formatting(sheet, 3, journal_df.columns.tolist())
        self._auto_adjust_columns(sheet, journal_df)
    
    def _create_summary_sheet(self, workbook: Workbook, transactions_df: pd.DataFrame, 
                            journal_entries: List[Dict[str, Any]], writer):
        """Create summary worksheet"""
        sheet = workbook.create_sheet("Summary", 2)
        
        # Add title
        sheet.merge_cells('A1:D1')
        sheet['A1'] = "FINANCIAL SUMMARY"
        sheet['A1'].font = Font(size=16, bold=True)
        sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Calculate summary statistics
        summary_data = self._calculate_summary_data(transactions_df, journal_entries)
        
        # Write summary data
        row_idx = 3
        sheet.cell(row=row_idx, column=1, value="Metric")
        sheet.cell(row=row_idx, column=2, value="Value")
        
        for key, value in summary_data.items():
            row_idx += 1
            sheet.cell(row=row_idx, column=1, value=key.replace('_', ' ').title())
            
            if isinstance(value, (int, float)):
                sheet.cell(row=row_idx, column=2, value=value)
                if 'amount' in key.lower() or 'total' in key.lower():
                    sheet.cell(row=row_idx, column=2).number_format = '#,##0.00'
            else:
                sheet.cell(row=row_idx, column=2, value=str(value))
        
        # Format summary
        self._format_summary_sheet(sheet, len(summary_data) + 3)
    
    def _create_analysis_sheet(self, workbook: Workbook, transactions_df: pd.DataFrame, 
                             journal_entries: List[Dict[str, Any]], writer):
        """Create analysis worksheet with charts"""
        if transactions_df.empty:
            return
        
        sheet = workbook.create_sheet("Analysis", 3)
        
        # Add title
        sheet.merge_cells('A1:D1')
        sheet['A1'] = "FINANCIAL ANALYSIS"
        sheet['A1'].font = Font(size=16, bold=True)
        sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Category analysis
        if 'category' in transactions_df.columns:
            category_totals = transactions_df.groupby('category')['amount'].sum().reset_index()
            category_totals = category_totals.sort_values('amount', ascending=False)
            
            # Write category analysis
            sheet.cell(row=3, column=1, value="Category")
            sheet.cell(row=3, column=2, value="Total Amount")
            
            for idx, (_, row) in enumerate(category_totals.iterrows(), 4):
                sheet.cell(row=idx, column=1, value=row['category'])
                sheet.cell(row=idx, column=2, value=row['amount'])
            
            # Create pie chart
            if len(category_totals) > 0:
                chart = PieChart()
                labels = Reference(sheet, min_col=1, min_row=4, max_row=3 + len(category_totals))
                data = Reference(sheet, min_col=2, min_row=3, max_row=3 + len(category_totals))
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(labels)
                chart.title = "Expense/Income by Category"
                sheet.add_chart(chart, "D3")
        
        # Format analysis sheet
        self._format_analysis_sheet(sheet)
    
    def _calculate_summary_data(self, transactions_df: pd.DataFrame, journal_entries: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Calculate summary statistics"""
        summary = {}
        
        if not transactions_df.empty:
            # Basic transaction stats
            summary['total_transactions'] = len(transactions_df)
            summary['total_credit_amount'] = transactions_df[transactions_df['type'] == 'CR']['amount'].sum()
            summary['total_debit_amount'] = transactions_df[transactions_df['type'] == 'DR']['amount'].sum()
            summary['net_balance'] = summary['total_credit_amount'] - summary['total_debit_amount']
            
            # Date range
            if 'date' in transactions_df.columns:
                try:
                    dates = pd.to_datetime(transactions_df['date'])
                    summary['start_date'] = dates.min().strftime('%d/%m/%Y')
                    summary['end_date'] = dates.max().strftime('%d/%m/%Y')
                except:
                    pass
            
            # Category stats
            if 'category' in transactions_df.columns:
                unique_categories = transactions_df['category'].nunique()
                summary['unique_categories'] = unique_categories
                most_common_category = transactions_df['category'].mode()
                if not most_common_category.empty:
                    summary['most_common_category'] = most_common_category.iloc[0]
        
        if journal_entries:
            summary['total_journal_entries'] = len(journal_entries)
        
        summary['report_generated'] = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        
        return summary
    
    def _format_data_range(self, sheet, start_row, start_col, end_row, end_col):
        """Apply formatting to data range"""
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)
                
                # Apply appropriate style based on content
                if row == start_row:  # Header row
                    self._apply_style(cell, self.styles['header'])
                elif isinstance(cell.value, (int, float)) and cell.value != 0:
                    if 'debit' in str(cell.value).lower() or (isinstance(cell.value, (int, float)) and cell.value < 0):
                        self._apply_style(cell, self.styles['debit'])
                    else:
                        self._apply_style(cell, self.styles['amount'])
                elif self._looks_like_date(cell.value):
                    self._apply_style(cell, self.styles['date'])
                else:
                    self._apply_style(cell, self.styles['text'])
    
    def _apply_header_formatting(self, sheet, header_row, headers):
        """Apply formatting to header row"""
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=header_row, column=col_idx)
            cell.value = header.replace('_', ' ').title()
            self._apply_style(cell, self.styles['header'])
    
    def _apply_style(self, cell, style):
        """Apply style dictionary to cell"""
        for attr, value in style.items():
            setattr(cell, attr, value)
    
    def _auto_adjust_columns(self, sheet, df):
        """Auto-adjust column widths based on content"""
        for col_idx, column in enumerate(df.columns, 1):
            max_length = 0
            column_letter = self._get_column_letter(col_idx)
            
            # Check header length
            max_length = max(max_length, len(str(column)))
            
            # Check data length
            for row in range(1, min(len(df) + 2, 1000)):  # Limit to first 1000 rows for performance
                try:
                    cell_value = sheet.cell(row=row, column=col_idx).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _get_column_letter(self, col_idx):
        """Convert column index to letter (1->A, 2->B, etc.)"""
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result
    
    def _add_totals_row(self, sheet, df, totals_row):
        """Add totals row to transactions"""
        if 'amount' in df.columns:
            total_amount = df['amount'].sum()
            sheet.cell(row=totals_row, column=1, value="TOTAL")
            sheet.cell(row=totals_row, column=df.columns.get_loc('amount') + 1, value=total_amount)
            self._apply_style(sheet.cell(row=totals_row, column=df.columns.get_loc('amount') + 1), self.styles['total'])
    
    def _format_summary_sheet(self, sheet, data_rows):
        """Format summary sheet"""
        # Apply formatting to summary data
        for row in range(3, data_rows + 1):
            for col in range(1, 3):
                cell = sheet.cell(row=row, column=col)
                if row == 3:  # Header
                    self._apply_style(cell, self.styles['header'])
                elif col == 2 and isinstance(cell.value, (int, float)):
                    self._apply_style(cell, self.styles['amount'])
                else:
                    self._apply_style(cell, self.styles['text'])
        
        # Auto-adjust columns
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 20
    
    def _format_analysis_sheet(self, sheet):
        """Format analysis sheet"""
        # Auto-adjust columns
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 20
    
    def _apply_auto_filters(self, workbook):
        """Apply auto-filters to data sheets"""
        for sheet_name in ['Transactions', 'Journal Entries']:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                # Find the data range (skip title rows)
                max_row = sheet.max_row
                max_col = sheet.max_column
                if max_row > 3 and max_col > 0:
                    sheet.auto_filter.ref = f"A3:{self._get_column_letter(max_col)}{max_row}"
    
    def _looks_like_date(self, value):
        """Check if value looks like a date"""
        if not value:
            return False
        date_indicators = ['/', '-', 'jan', 'feb', 'mar', 'apr', 'may', 'jun', 
                          'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
        value_str = str(value).lower()
        return any(indicator in value_str for indicator in date_indicators)
    
    def export_to_excel_buffer(self, transactions_df: pd.DataFrame, journal_entries: List[Dict[str, Any]]) -> bytes:
        """Export to Excel and return as bytes buffer"""
        import tempfile
        import io
        
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                self.export_to_excel(transactions_df, journal_entries, tmp_file.name, include_charts=False)
                
                # Read the file back into memory
                with open(tmp_file.name, 'rb') as f:
                    excel_buffer = io.BytesIO(f.read())
                
                # Clean up
                os.unlink(tmp_file.name)
                
                return excel_buffer.getvalue()
                
        except Exception as e:
            print(f"Error creating Excel buffer: {e}")
            raise